import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Fetch problem details from Codeforces
def fetch_problem_details(problem_links):
    data = []
    for link in problem_links:
        try:
            parts = link.split("/")
            contest_id = parts[-2]
            problem_index = parts[-1]
            
            # Fetch data from API
            response = requests.get("https://codeforces.com/api/problemset.problems")
            response.raise_for_status()
            problems = response.json().get("result", {}).get("problems", [])
            
            # Find the specific problem
            for problem in problems:
                if str(problem.get("contestId")) == contest_id and problem.get("index") == problem_index:
                    data.append({
                        "Problem Name": problem.get("name", "Unknown"),
                        "Problem Link": link,
                        "Rating": problem.get("rating", "Unknown"),
                        "Tags": ", ".join(problem.get("tags", [])),
                        "Status": "Not Solved",  # Default
                        "Favorite": "No"  # Default
                    })
                    break
        except Exception as e:
            print(f"Error fetching details for {link}: {e}")
    return data

# Add colors based on cell values
def apply_styles(ws):
    status_colors = {
        "Attempted": "FFFF00",  # Yellow
        "Solved": "00FF00",     # Green
        "Unsolved": "FF0000",   # Red
    }
    favorite_color = "ADD8E6"  # Light Blue for favorites

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=6):
        # Status column
        status_cell = row[0]
        if status_cell.value in status_colors:
            color = status_colors[status_cell.value]
            status_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Favorite column
        favorite_cell = row[1]
        if favorite_cell.value == "Yes":
            for cell in ws[row[0].row]:
                cell.fill = PatternFill(start_color=favorite_color, end_color=favorite_color, fill_type="solid")

# Save the Excel file
def create_excel(data, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Codeforces Problems"

    # Headers
    headers = ["Problem Name", "Problem Link", "Rating", "Tags", "Status", "Favorite"]
    ws.append(headers)

    # Style headers
    for col in ws.iter_cols(min_row=1, max_row=1, max_col=len(headers)):
        for cell in col:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows
    for item in data:
        ws.append(list(item.values()))

    # Auto-size columns
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 2

    # Apply styles for Status and Favorite columns
    apply_styles(ws)

    # Save the file
    wb.save(filename)
    print(f"Excel file created: {filename}")

# Input problem links
problem_links = [
    "https://codeforces.com/problemset/problem/2047/A",
    "https://codeforces.com/problemset/problem/2042/B",
    "https://codeforces.com/problemset/problem/2042/A",
    "https://codeforces.com/problemset/problem/2039/B",
    "https://codeforces.com/problemset/problem/2039/A",
    "https://codeforces.com/problemset/problem/2037/C",
    "https://codeforces.com/problemset/problem/2035/B",
    "https://codeforces.com/problemset/problem/2031/B",
    "https://codeforces.com/problemset/problem/2030/B",
    "https://codeforces.com/problemset/problem/2028/A",
    "https://codeforces.com/problemset/problem/2025/A",
    "https://codeforces.com/problemset/problem/2021/A",
    "https://codeforces.com/problemset/problem/2002/B"
]

# Fetch problem details
problems_data = fetch_problem_details(problem_links)

# Create Excel file
create_excel(problems_data, "Codeforces_Problems_Automated.xlsx")
